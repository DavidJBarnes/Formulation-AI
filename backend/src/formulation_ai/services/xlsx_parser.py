from __future__ import annotations

import re
from dataclasses import dataclass, field
from io import BytesIO

import openpyxl


@dataclass
class ParsedIngredient:
    name: str
    unit: str | None


@dataclass
class ParsedProperty:
    name: str
    unit: str | None


@dataclass
class ParsedProduct:
    label: str
    ingredients: dict[str, float] = field(default_factory=dict)
    properties: dict[str, float] = field(default_factory=dict)


@dataclass
class ParsedTarget:
    property_name: str
    goal: str
    reference_label: str | None


@dataclass
class ParsedUpload:
    ingredients: list[ParsedIngredient]
    properties: list[ParsedProperty]
    base_products: list[ParsedProduct]
    targets: list[ParsedTarget]


def parse_xlsx(data: bytes) -> ParsedUpload:
    wb = openpyxl.load_workbook(BytesIO(data), read_only=True, data_only=True)

    if "Products" not in wb.sheetnames:
        raise ValueError("XLSX must contain a 'Products' sheet")
    if "Targets" not in wb.sheetnames:
        raise ValueError("XLSX must contain a 'Targets' sheet")

    ws_products = wb["Products"]
    ws_targets = wb["Targets"]

    # --- Products sheet ---
    rows = list(ws_products.iter_rows(values_only=True))
    if not rows:
        raise ValueError("Products sheet is empty")

    headers = [str(h).strip() if h is not None else "" for h in rows[0]]

    ingredient_cols: list[tuple[int, str, str | None]] = []
    property_cols: list[tuple[int, str, str | None]] = []

    for i, h in enumerate(headers):
        if i == 0:
            continue
        if h.startswith("Ingredient:"):
            name, unit = _parse_col_header(h[len("Ingredient:"):].strip())
            ingredient_cols.append((i, name, unit))
        elif h.startswith("Property:"):
            name, unit = _parse_col_header(h[len("Property:"):].strip())
            property_cols.append((i, name, unit))

    base_products: list[ParsedProduct] = []
    for row in rows[1:]:
        if not row or row[0] is None:
            continue
        label = str(row[0]).strip()
        if not label:
            continue
        product = ParsedProduct(label=label)
        for idx, name, _ in ingredient_cols:
            val = row[idx] if idx < len(row) else None
            if val is not None:
                f = _to_float(val)
                if f is not None:
                    product.ingredients[name] = f
        for idx, name, _ in property_cols:
            val = row[idx] if idx < len(row) else None
            if val is not None:
                f = _to_float(val)
                if f is not None:
                    product.properties[name] = f
        base_products.append(product)

    # --- Targets sheet ---
    t_rows = list(ws_targets.iter_rows(values_only=True))
    targets: list[ParsedTarget] = []

    if not t_rows:
        return ParsedUpload(
            ingredients=[ParsedIngredient(n, u) for _, n, u in ingredient_cols],
            properties=[ParsedProperty(n, u) for _, n, u in property_cols],
            base_products=base_products,
            targets=[],
        )

    t_hdrs = [str(h).strip().lower() if h is not None else "" for h in t_rows[0]]
    prop_col = _find_col(t_hdrs, ["property"])
    goal_col = _find_col(t_hdrs, ["goal"])
    ref_col = _find_col(t_hdrs, ["reference", "ref"])

    for row in t_rows[1:]:
        if not row or _cell(row, prop_col) is None:
            continue
        prop_name = str(_cell(row, prop_col)).strip()
        raw_goal = _cell(row, goal_col)
        if raw_goal is None:
            continue
        goal = str(raw_goal).strip()
        raw_ref = _cell(row, ref_col)
        ref = str(raw_ref).strip() if raw_ref is not None else None
        if ref and ref.lower() in ("absolute", "none", "-", "n/a", ""):
            ref = None
        targets.append(ParsedTarget(property_name=prop_name, goal=goal, reference_label=ref))

    return ParsedUpload(
        ingredients=[ParsedIngredient(n, u) for _, n, u in ingredient_cols],
        properties=[ParsedProperty(n, u) for _, n, u in property_cols],
        base_products=base_products,
        targets=targets,
    )


def _parse_col_header(s: str) -> tuple[str, str | None]:
    """'Name (unit)' → (name, unit).  'Name' → (name, None)."""
    m = re.match(r"^(.+?)\s*\(([^)]*)\)\s*$", s)
    if m:
        unit = m.group(2).strip() or None
        return m.group(1).strip(), unit
    return s.strip(), None


def _to_float(val: object) -> float | None:
    try:
        return float(val)  # type: ignore[arg-type]
    except (TypeError, ValueError):
        return None


def _find_col(headers: list[str], names: list[str]) -> int:
    for name in names:
        for i, h in enumerate(headers):
            if h == name:
                return i
    return 0


def _cell(row: tuple, idx: int):
    return row[idx] if idx < len(row) else None
