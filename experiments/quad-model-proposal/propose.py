"""LLM-only proposal generator for the Quad Model audition.

Reads Jun's Quad Model.xlsx (9 rows of Input 1-3 x Output 1-3 observations),
asks Claude Opus 4.7 to propose K new candidate input vectors that satisfy a
target, and writes the candidates + predicted (value, sigma) per output to
proposals.xlsx — same shape as Jun's "Sheet1 predictions" but with novel
candidates rather than re-predictions.

Usage:
    ANTHROPIC_API_KEY=... uv run python propose.py
    ANTHROPIC_API_KEY=... uv run python propose.py --target "Minimize Output 1 magnitude; Output 3 near 20" --n 8
"""

from __future__ import annotations

import argparse
import json
import re
from pathlib import Path

import anthropic
import openpyxl

SCRIPT_DIR = Path(__file__).resolve().parent
DEFAULT_INPUT = SCRIPT_DIR / "Quad Model.xlsx"
DEFAULT_OUTPUT = SCRIPT_DIR / "proposals.xlsx"
DEFAULT_TARGET = (
    "Maximize Output 2 while keeping |Output 1| < 500 and Output 3 within [35, 45]."
)
DEFAULT_N = 5
MODEL = "claude-opus-4-7"

OUTPUT_COLUMNS = [
    "Candidate",
    "Input 1",
    "Input 2",
    "Input 3",
    "Pred Output 1",
    "Pred Output 2",
    "Pred Output 3",
    "σ Output 1",
    "σ Output 2",
    "σ Output 3",
    "Rationale",
]


def load_observations(path: Path) -> list[dict]:
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["Actual Model"]
    rows = list(ws.iter_rows(values_only=True))
    header = rows[0]
    return [dict(zip(header, r, strict=True)) for r in rows[1:]]


def build_prompt(observations: list[dict], target: str, n: int) -> str:
    table = "\n".join(
        f"{r['ROW NUM']:>3} | {r['Input 1']:>6} | {r['Input 2']:>6} | {r['Input 3']:>6} | "
        f"{r['Output 1']:>8} | {r['Output 2']:>8} | {r['Output 3']:>6}"
        for r in observations
    )
    return f"""You are the proposal engine for a closed-loop formulation optimization system.

A small set of multivariate observations has been collected. Given the 9 observations \
below and a user-specified target, propose {n} NEW candidate input vectors that you \
predict will satisfy the target.

Historical observations:
ROW | Input 1 | Input 2 | Input 3 | Output 1 | Output 2 | Output 3
{table}

Target: {target}

Rules:
- Each candidate must be a novel (Input 1, Input 2, Input 3) vector — do NOT repeat any historical row.
- Keep inputs plausible: within or near the observed range.
- For each candidate, emit predicted values for all three outputs AND a 1-sigma \
uncertainty per output. With only 9 training points, uncertainty should generally \
be non-trivial — be honest rather than overconfident.
- Rank candidates by how confidently you expect them to meet the target (candidate 1 = best).
- Include a 1-2 sentence rationale for each candidate — what signal from the historical \
data informs this choice.

Respond with a single JSON object (no prose before or after) matching this schema exactly:

{{
  "candidates": [
    {{
      "input_1": number,
      "input_2": number,
      "input_3": number,
      "predicted_output_1": number,
      "predicted_output_2": number,
      "predicted_output_3": number,
      "sigma_output_1": number,
      "sigma_output_2": number,
      "sigma_output_3": number,
      "rationale": "string"
    }}
  ]
}}
"""


def call_claude(prompt: str) -> str:
    client = anthropic.Anthropic()
    msg = client.messages.create(
        model=MODEL,
        max_tokens=4000,
        messages=[{"role": "user", "content": prompt}],
    )
    parts = [b.text for b in msg.content if getattr(b, "type", None) == "text"]
    return "".join(parts)


def parse_response(text: str) -> dict:
    match = re.search(r"\{.*\}", text, re.DOTALL)
    if not match:
        raise ValueError(f"No JSON found in LLM response: {text[:400]!r}")
    return json.loads(match.group(0))


def write_xlsx(result: dict, target: str, out_path: Path) -> int:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "LLM Proposals"

    ws.append(["Target", target])
    ws.append(["Model", MODEL])
    ws.append([])
    ws.append(OUTPUT_COLUMNS)

    for i, c in enumerate(result["candidates"], 1):
        ws.append(
            [
                i,
                c["input_1"],
                c["input_2"],
                c["input_3"],
                c["predicted_output_1"],
                c["predicted_output_2"],
                c["predicted_output_3"],
                c["sigma_output_1"],
                c["sigma_output_2"],
                c["sigma_output_3"],
                c["rationale"],
            ]
        )

    widths = [11, 9, 9, 9, 14, 14, 14, 11, 11, 11, 80]
    for col_idx, width in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width

    wb.save(out_path)
    return len(result["candidates"])


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--input", type=Path, default=DEFAULT_INPUT)
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT)
    parser.add_argument("--target", default=DEFAULT_TARGET)
    parser.add_argument("--n", type=int, default=DEFAULT_N)
    args = parser.parse_args()

    observations = load_observations(args.input)
    prompt = build_prompt(observations, args.target, args.n)
    text = call_claude(prompt)
    result = parse_response(text)
    count = write_xlsx(result, args.target, args.output)
    print(f"Wrote {count} candidates to {args.output}")


if __name__ == "__main__":
    main()
