"""Generate the Formulation-AI upload XLSX template + a worked paint example.

Two outputs land next to this script:
  - template.xlsx       : empty, ready to fill in
  - paint-example.xlsx  : populated with Jun's paint scenario from the
                          original 2026-04-23 email thread

Run:
    python generate.py
"""

from __future__ import annotations

from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

SCRIPT_DIR = Path(__file__).resolve().parent

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)
TITLE_FONT = Font(bold=True, size=14)
SECTION_FONT = Font(bold=True, size=11)


# ---- paint-example content ------------------------------------------------

INGREDIENTS: list[tuple[str, str]] = [
    ("Titanium dioxide", "% w/w"),
    ("Calcium carbonate", "% w/w"),
    ("Acrylic resin", "% w/w"),
    ("Water", "% w/w"),
    ("Surfactant", "% w/w"),
    ("Defoamer", "% w/w"),
    ("Thickener", "% w/w"),
]

PROPERTIES: list[tuple[str, str]] = [
    ("Viscosity", "KU"),
    ("Opacity", "%"),
    ("Adhesion", "MPa"),
    ("Durability", "years"),
]

PAINT_PRODUCTS: list[tuple[str, list[float], list[float]]] = [
    ("Paint A", [22.0, 18.0, 14.5, 38.0, 1.2, 0.8, 0.5], [102, 95, 4.2, 8]),
    ("Paint B", [12.0, 28.0, 11.5, 42.0, 1.0, 0.7, 0.6], [98, 82, 3.8, 6]),
    ("Paint C", [20.0, 16.0, 18.0, 35.0, 1.4, 0.9, 0.5], [110, 92, 4.8, 12]),
    ("Paint D", [14.0, 24.0, 14.0, 40.0, 1.1, 0.8, 0.7], [104, 80, 4.0, 9]),
]

PAINT_TARGETS: list[tuple[str, str, str, str]] = [
    ("Viscosity", "=105", "absolute", "Hit a viscosity target."),
    ("Opacity", "+10%", "average of base", "10% better than current product line."),
    ("Adhesion", ">=4.5", "absolute", "Industry-grade adhesion floor."),
    ("Durability", "+10%", "average of base", "10% longer life than current product line."),
]


# ---- sheet builders -------------------------------------------------------

def write_readme(ws) -> None:
    ws.title = "README"
    lines: list[tuple[str, Font | None]] = [
        ("Formulation-AI — upload template", TITLE_FONT),
        ("", None),
        ("Three sheets:", SECTION_FONT),
        ("  1. README — this sheet (read once, ignore later).", None),
        ("  2. Products — your existing formulations + their measured properties.", None),
        ("  3. Targets — optimization targets for the new project.", None),
        ("", None),
        ("Products sheet conventions", SECTION_FONT),
        ("  • One row per product. First column is the product ID/name.", None),
        ("  • Ingredient columns are named:   Ingredient: <name> (<unit>)", None),
        ("       e.g.  Ingredient: Titanium dioxide (% w/w)", None),
        ("  • Property columns are named:    Property: <name> (<unit>)", None),
        ("       e.g.  Property: Viscosity (KU)", None),
        ("  • Leave ingredient cells blank if the product does not contain that ingredient.", None),
        ("  • Property cells should be filled for every product (these are measured outcomes).", None),
        ("", None),
        ("Targets sheet conventions", SECTION_FONT),
        ("  • One row per target. Property name must match a Property column on the Products sheet.", None),
        ("  • Goal column accepts these operators:", None),
        ("       =N        Hit absolute value N", None),
        ("       >=N       Minimum N (must be at least N)", None),
        ("       <=N       Maximum N (must be at most N)", None),
        ("       +N%       Increase by N% (vs. Reference)", None),
        ("       -N%       Decrease by N% (vs. Reference)", None),
        ("       [a,b]     Stay within range [a, b]", None),
        ("  • Reference column (only needed for +%/-%): which product or aggregate to compare to.", None),
        ("       e.g.  Paint A   |   average of base", None),
        ("", None),
        ("Tips", SECTION_FONT),
        ("  • Keep ingredient and property names consistent across products.", None),
        ("  • Units are part of the column name — change the unit, change the column.", None),
        ("  • See paint-example.xlsx for a fully filled-in version using the original Jun scenario.", None),
    ]
    for r, (text, font) in enumerate(lines, 1):
        cell = ws.cell(row=r, column=1, value=text)
        if font is not None:
            cell.font = font
    ws.column_dimensions["A"].width = 95


def write_products(ws, products: list[tuple[str, list[float], list[float]]] | None) -> None:
    ws.title = "Products"
    headers = ["Product"]
    headers.extend(f"Ingredient: {name} ({unit})" for name, unit in INGREDIENTS)
    headers.extend(f"Property: {name} ({unit})" for name, unit in PROPERTIES)
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(wrap_text=True, vertical="center")
        ws.column_dimensions[get_column_letter(col_idx)].width = 22
    ws.row_dimensions[1].height = 36
    ws.column_dimensions["A"].width = 14
    ws.freeze_panes = "B2"
    if not products:
        return
    for r, (product_id, ingredient_values, property_values) in enumerate(products, 2):
        ws.cell(row=r, column=1, value=product_id)
        for c, value in enumerate(ingredient_values, 2):
            ws.cell(row=r, column=c, value=value)
        offset = 2 + len(INGREDIENTS)
        for c, value in enumerate(property_values, offset):
            ws.cell(row=r, column=c, value=value)


def write_targets(ws, targets: list[tuple[str, str, str, str]] | None) -> None:
    ws.title = "Targets"
    headers = [("Property", 22), ("Goal", 14), ("Reference", 22), ("Notes", 60)]
    for col_idx, (label, width) in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.freeze_panes = "A2"
    if not targets:
        return
    for r, row in enumerate(targets, 2):
        for c, value in enumerate(row, 1):
            ws.cell(row=r, column=c, value=value)


def build(
    output: Path,
    *,
    products: list[tuple[str, list[float], list[float]]] | None = None,
    targets: list[tuple[str, str, str, str]] | None = None,
) -> None:
    wb = openpyxl.Workbook()
    write_readme(wb.active)
    write_products(wb.create_sheet("Products"), products)
    write_targets(wb.create_sheet("Targets"), targets)
    wb.save(output)


def main() -> None:
    template = SCRIPT_DIR / "template.xlsx"
    example = SCRIPT_DIR / "paint-example.xlsx"
    build(template)
    build(example, products=PAINT_PRODUCTS, targets=PAINT_TARGETS)
    print(f"Wrote {template}")
    print(f"Wrote {example}")


if __name__ == "__main__":
    main()
